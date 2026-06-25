"""Parse `Lộ trình.xlsx` → CSV audit + (optional) INSERT into app.tProductionCurve.

Usage:
    # Step 1 — parse + xuất audit CSV (mặc định, KHÔNG đụng DB)
    python -X utf8 scripts/seed_curve.py --xlsx "C:\\Users\\OS\\Downloads\\Lộ trình.xlsx"

    # Step 2 — sau khi rà xong CSV, chạy lại với --apply để insert DB
    python -X utf8 scripts/seed_curve.py --xlsx "..." --apply

Logic phân loại (theo cột `Phân loại ĐH` trong Excel):
    "Đặc biệt-CĐ1" → Category="Đặc biệt", NDSXLevel=1
    "Đặc biệt-CĐ2" → ..., NDSXLevel=2
    "Đặc biệt-CĐ3" → ..., NDSXLevel=3
    "Mới-CĐx", "Lặp lại-CĐx" → tương tự
    "Vest (Áo, Quần)" → Category="Vest", NDSXLevel=0

Logic parse ratio:
    "0,16" / "0,234" → 0.16 / 0.234   (string Việt format)
    1105 → 1.105                       (lỗi format Excel mất dấu phẩy, chia 1000)
    1.05 → 1.05                        (bình thường)
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Iterable

import openpyxl

# Mapping từ "Phân loại ĐH" trong Excel → (Category, NDSXLevel)
CATEGORY_RX = re.compile(r"^(Đặc biệt|Mới|Lặp lại)\s*-\s*CĐ(\d)$")
VEST_RX = re.compile(r"^Vest", re.IGNORECASE)

CURVE_TYPE_KEEP = "Hiệu suất RC"  # chỉ seed đường này; bỏ "Bình quân ngày SX"


def parse_label(raw: str | None) -> tuple[str, int] | None:
    """Map cell value in col A → (Category, NDSXLevel). None = skip row."""
    if not raw or not isinstance(raw, str):
        return None
    s = raw.strip()
    m = CATEGORY_RX.match(s)
    if m:
        return m.group(1), int(m.group(2))
    if VEST_RX.match(s):
        return "Vest", 0
    return None


def parse_ratio(v) -> float | None:
    """Parse a single ratio cell handling Excel format quirks.

    Returns None on empty/invalid; raises ValueError on unparseable strings.
    """
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        # Lỗi format mất dấu phẩy: số nguyên 3-4 chữ số mà giá trị > 10
        # → đáng lẽ là decimal (1105 → 1.105). Chia 1000.
        if f > 10:
            return round(f / 1000.0, 5)
        return round(f, 5)
    if isinstance(v, str):
        s = v.strip().replace(",", ".")
        if not s:
            return None
        try:
            f = float(s)
        except ValueError as e:
            raise ValueError(f"Cannot parse ratio: {v!r}") from e
        if f > 10:
            return round(f / 1000.0, 5)
        return round(f, 5)
    raise ValueError(f"Unsupported ratio type: {type(v).__name__} = {v!r}")


def extract_rows(xlsx_path: Path) -> tuple[list[dict], list[dict]]:
    """Return (curve_rows, audit_anomalies).

    curve_rows: list of dict ready to insert into app.tProductionCurve
    audit_anomalies: list of dict for rows where format-fix kicked in
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active  # "Lộ trình HS"

    curve_rows: list[dict] = []
    anomalies: list[dict] = []

    # Hàng 1 = title; hàng 2 = header. Day numbers nằm ở cột E..GB (5..184)
    # Row 2 col 5..184 = 1..180
    DAY_COL_START = 5
    DAY_COL_END = ws.max_column  # auto-detect

    for row_idx in range(3, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        curve_type_raw = ws.cell(row=row_idx, column=4).value
        parsed_label = parse_label(label)
        if parsed_label is None:
            continue
        if curve_type_raw != CURVE_TYPE_KEEP:
            continue  # bỏ Bình quân ngày SX và các loại khác

        category, ndsx = parsed_label

        for col_idx in range(DAY_COL_START, DAY_COL_END + 1):
            day_n_raw = ws.cell(row=2, column=col_idx).value
            if not isinstance(day_n_raw, int):
                continue
            day_n = day_n_raw
            raw_val = ws.cell(row=row_idx, column=col_idx).value
            if raw_val is None or raw_val == "":
                continue
            try:
                ratio = parse_ratio(raw_val)
            except ValueError as e:
                anomalies.append({
                    "row": row_idx, "col": col_idx, "day": day_n,
                    "category": category, "ndsx": ndsx,
                    "raw": repr(raw_val),
                    "issue": str(e),
                })
                continue
            if ratio is None:
                continue

            curve_rows.append({
                "category": category,
                "ndsx_level": ndsx,
                "day_n": day_n,
                "ratio": ratio,
                "raw": raw_val,
            })

            # Flag nếu lỗi format (raw là numeric > 10)
            if isinstance(raw_val, (int, float)) and float(raw_val) > 10:
                anomalies.append({
                    "row": row_idx, "col": col_idx, "day": day_n,
                    "category": category, "ndsx": ndsx,
                    "raw": raw_val,
                    "fixed_to": ratio,
                    "issue": "comma-decimal lost — divided by 1000",
                })

    return curve_rows, anomalies


def write_csv(rows: Iterable[dict], path: Path, fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fieldnames})


def apply_to_db(rows: list[dict]) -> None:
    """Bulk delete + insert into app.tProductionCurve. Requires app.db importable."""
    import os
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from app.db import get_conn  # noqa: E402

    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM app.tProductionCurve")
        cur.fast_executemany = True
        cur.executemany(
            "INSERT INTO app.tProductionCurve "
            "(Category, NDSXLevel, DayN, Ratio) VALUES (?,?,?,?)",
            [(r["category"], r["ndsx_level"], r["day_n"], r["ratio"])
             for r in rows],
        )
    print(f"✓ Inserted {len(rows)} rows into app.tProductionCurve")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, type=Path,
                    help="Path tới file Lộ trình.xlsx")
    ap.add_argument("--out-dir", type=Path,
                    default=Path(__file__).resolve().parent / "out",
                    help="Thư mục xuất CSV audit (mặc định scripts/out/)")
    ap.add_argument("--apply", action="store_true",
                    help="Sau khi parse, INSERT vào app.tProductionCurve")
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f"✗ File not found: {args.xlsx}", file=sys.stderr)
        return 1

    args.out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Parsing {args.xlsx} ...")
    rows, anomalies = extract_rows(args.xlsx)
    print(f"  → {len(rows)} curve rows, {len(anomalies)} anomalies flagged")

    curve_csv = args.out_dir / "curve_seed.csv"
    audit_csv = args.out_dir / "curve_audit.csv"

    write_csv(
        rows, curve_csv,
        ["category", "ndsx_level", "day_n", "ratio", "raw"],
    )
    write_csv(
        anomalies, audit_csv,
        ["row", "col", "day", "category", "ndsx",
         "raw", "fixed_to", "issue"],
    )

    print(f"  → wrote {curve_csv}")
    print(f"  → wrote {audit_csv}")

    # Sample print: anomalies head
    if anomalies:
        print("\nFirst 5 anomalies (raw → fixed):")
        for a in anomalies[:5]:
            print(f"  D{a['day']:>3} | {a['category']:>10}/CĐ{a['ndsx']} | "
                  f"{a['raw']!r} → {a.get('fixed_to', '?')}")

    if args.apply:
        print("\nApplying to DB ...")
        apply_to_db(rows)

    return 0


if __name__ == "__main__":
    sys.exit(main())
